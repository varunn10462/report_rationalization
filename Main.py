import re
import pandas as pd
import random
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# File paths (replace with your actual paths)
twb_file = "C:\\Users\\varunsharma\\Desktop\\Sanu_task_version_2\\final doc\\ADS_Channel_Summary_Weekly.twb"  # Input Tableau workbook
output_excel = "updated_sensitive_data_detection.xlsx"  # Output Excel file

# Step 1: Parse the `.twb` file to extract SQL queries
def extract_sql_from_twb(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    sql_queries = []

    for datasource in root.findall(".//datasource"):
        for connection in datasource.findall(".//connection"):
            custom_sql = connection.find(".//relation")
            if custom_sql is not None and custom_sql.attrib.get("type") == "text":
                sql_queries.append(custom_sql.text)
    
    return sql_queries

# Step 2: Detect sensitive data in SQL queries
def detect_sensitive_data(sql_queries):
    patterns = {
        "Name": r"(?:[A-Z][a-z]+\s+[A-Z][a-z]+)",
        "Account Number": r"\b\d{16}\b",
        "Address": r"(?:\d{1,5}\s\w+\s\w+,\s\w+,\s\w+, \w{2},\s\d{5})",
        "Member": r"(?:Premium Member|Regular Member|Basic Member)",
        "DOB": r"\d{4}-\d{2}-\d{2}",
        "Gender": r"(?:Male|Female)",
        "Phone number": r"\b\d{3}-\d{3}-\d{4}\b",
        "Email": r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}",
        "Card Number": r"\b\d{16}\b",
        "SSN": r"\b\d{3}-\d{2}-\d{4}\b",
        "TIN": r"\b\d{2}-\d{7}\b"
    }

    results = []
    for sql_query in sql_queries:
        for name, pattern in patterns.items():
            matches = re.findall(pattern, sql_query)
            for match in matches:
                results.append({
                    "Source": "Extracted SQL",
                    "Report name": "Sensitive Data Detection",
                    "DB": "DatabaseName",
                    "Schema": "SchemaName",
                    "Table": "TableName",
                    "Is Sensitive": random.choice(["Yes", "No"]),  # Randomize Yes/No
                    name: match
                })
    return results

# Step 3: Write results to an Excel file and format it
def save_to_excel(results, output_file):
    df = pd.DataFrame(results)
    df.to_excel(output_file, index=False)

    # Load the workbook and select the active worksheet
    wb = load_workbook(output_file)
    ws = wb.active

    # Apply formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(output_file)
    print(f"Results saved to {output_file}")

# Main execution flow
if __name__ == "__main__":
    print("Extracting SQL queries from Tableau workbook...")
    sql_queries = extract_sql_from_twb(twb_file)
    
    if not sql_queries:
        print("No SQL queries found in the Tableau workbook.")
    else:
        print(f"Found {len(sql_queries)} SQL query(ies). Detecting sensitive data...")
        results = detect_sensitive_data(sql_queries)
        save_to_excel(results, output_excel)
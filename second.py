import re
import pandas as pd
import random
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

# Directory containing the .twb files
twb_directory = "C:\\Users\\varunsharma\\Desktop\\Sanu_task_version_2\\final doc"
output_excel = "updated_sensitive_data_detection.xlsx"  # Output Excel file

# Step 1: Parse the `.twb` file to extract SQL queries
def extract_sql_from_twb(file_path):
    try:
        # Parse the XML content of the .twb file
        tree = ET.parse(file_path)
        root = tree.getroot()
        sql_queries = []

        # Find all datasource elements
        for datasource in root.findall(".//datasource"):
            # Find all connection elements within each datasource
            for connection in datasource.findall(".//connection"):
                # Find the custom SQL relation element
                custom_sql = connection.find(".//relation")
                if custom_sql is not None and custom_sql.attrib.get("type") == "text":
                    # Append the SQL query text to the list
                    sql_queries.append(custom_sql.text)
        
        return sql_queries
    except ET.ParseError:
        print("Error parsing XML file: {}".format(file_path))
        return []

# Step 2: Detect sensitive data in SQL queries
def detect_sensitive_data(sql_queries):
    # Define patterns for detecting sensitive data
    patterns = {
        "Name": r"(?:[A-Z][a-z]+\s+[A-Z][a-z]+)",
        "Account Number": r"\b\d{16}\b",
        "Address": r"(?:\d{1,5}\s\w+\s\w+,\s\w+,\s\w+, \w{2},\s\d{5})",
        "Member": r"(?:Premium Member|Regular Member|Basic Member)",
        "DOB": r"\d{4}-\d{2}-\d{2}",
        "Gender": r"(?:Male|Female)",
        "Phone number": r"\b\d{3}-\d{3}-\d{4}\b",
        "Email": r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}",
        "Card Number": r"\b\d{16}\b",
        "SSN": r"\b\d{3}-\d{2}-\d{4}\b",
        "TIN": r"\b\d{2}-\d{7}\b"
    }

    results = []
    # Iterate over each SQL query
    for sql_query in sql_queries:
        # Check each pattern for matches in the SQL query
        for name, pattern in patterns.items():
            matches = re.findall(pattern, sql_query)
            for match in matches:
                # Append the detected sensitive data to the results list
                results.append({
                    "Source": "Extracted SQL",
                    "Report name": "Sensitive Data Detection",
                    "DB": "DatabaseName",
                    "Schema": "SchemaName",
                    "Table": "TableName",
                    "Is Sensitive": random.choice(["Yes", "No"]),  # Randomize Yes/No
                    name: match
                })
    return results

# Step 3: Write results to an Excel file and format it
def save_to_excel(results, output_file):
    # Convert the results list to a DataFrame
    df = pd.DataFrame(results)
    # Save the DataFrame to an Excel file
    df.to_excel(output_file, index=False)

    # Load the workbook and select the active worksheet
    wb = load_workbook(output_file)
    ws = wb.active

    # Apply formatting to the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Apply alignment to all cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment

    # Adjust column widths based on the maximum length of the content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the formatted workbook
    wb.save(output_file)
    print("Results saved to {}".format(output_file))

# Main execution flow
if __name__ == "__main__":
    all_results = []
    print("Scanning .twb files in directory...")

    # Iterate over all files in the specified directory
    for filename in os.listdir(twb_directory):
        if filename.endswith(".twb"):
            file_path = os.path.join(twb_directory, filename)
            print("Extracting SQL queries from {}...".format(filename))
            sql_queries = extract_sql_from_twb(file_path)
            
            if not sql_queries:
                print("No SQL queries found in {}.".format(filename))
            else:
                print("Found {} SQL query(ies) in {}. Detecting sensitive data...".format(len(sql_queries), filename))
                results = detect_sensitive_data(sql_queries)
                all_results.extend(results)
    
    # Save all results to the Excel file
    if all_results:
        save_to_excel(all_results, output_excel)
    else:
        print("No SQL queries found in any .twb files.")